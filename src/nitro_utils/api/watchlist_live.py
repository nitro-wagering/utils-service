"""Live watchlist endpoint — three-source composition (frozen S3 + live DB + recompute)."""

import io
import logging
from datetime import UTC, date, datetime

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from nitro_utils.database import get_db_session
from nitro_utils.date_utils import brisbane_today
from nitro_utils.live_data import fetch_live_odds, fetch_race_statuses, fetch_results
from nitro_utils.models import Track, UserBet
from nitro_utils.paper_ledger import fetch_paper_bets
from nitro_utils.s3_loader import load_frozen_watchlist

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/watchlist", tags=["watchlist"])


class WatchlistEntry(BaseModel):
    track: str
    country: str
    track_state: str | None
    race_number: int
    race_time: str
    race_name: str
    distance_m: int
    class_description: str
    track_condition: str
    horse: str
    jockey_name: str | None
    weight_kg: float | None
    barrier: int | None
    our_win: float
    win_pct: float
    win_trigger: float
    our_place: float
    place_pct: float
    place_trigger: float
    win_overlay_pct: float | None
    win_distance_to_trigger: float | None
    market_win: float | None
    market_place: float | None
    market_rank: int | None
    place_overlay_pct: float | None
    place_distance_to_trigger: float | None
    neds_win: float | None
    neds_place: float | None
    class_rank: int
    sim_order: int | None
    sim_win_pct: float | None
    sim_win_rank: int | None
    sim_place_pct: float | None
    sim_place_rank: int | None
    ml_win_pct: float | None
    ml_win_rank: int | None
    ml_place_pct: float | None
    ml_place_rank: int | None
    pf_time_rank: int | None
    neds_url: str | None
    in_monitor_net: str
    placed: str
    race_id: int
    race_date: str
    form_id: int
    horse_id: int
    actual_position: int | None
    actual_margin: float | None
    is_scratched: bool
    race_status: str | None
    bet_placed: bool
    bet_id: int | None
    bet_type: str | None
    odds_taken: float | None
    stake_aud: float | None
    result_position: int | None
    payout_aud: float | None
    profit_aud: float | None
    roi_pct: float | None
    paper_win_placed: bool
    paper_place_placed: bool
    paper_win_result: str | None
    paper_place_result: str | None
    paper_win_stake: float | None
    paper_place_stake: float | None
    paper_win_odds: float | None
    paper_place_odds: float | None


class WatchlistResponse(BaseModel):
    generated_at: str
    entry_count: int
    entries: list[WatchlistEntry]
    message: str | None = None


@router.get("", response_model=WatchlistResponse)
async def get_watchlist_live(
    date_str: str | None = Query(
        None, alias="date", description="Target date YYYY-MM-DD (default Brisbane today)"
    ),
    username: str | None = Query(None, description="Username to load bets for"),
    session: AsyncSession = Depends(get_db_session),
) -> WatchlistResponse:
    """Serve live watchlist: frozen predictions + live odds + live results.

    Three-source composition:
    1. Frozen predictions from S3: ml-v3/watchlists/{date}.csv
    2. Live odds: odds_snapshots DISTINCT ON latest by polled_at DESC
    3. Live results: race_entries.position + margin, races.race_status

    Recomputes overlays based on live market × frozen fair prices.
    Keeps in_monitor_net frozen (monitor's BET verdict at decision time).
    """
    # Parse target date
    if date_str:
        try:
            target_date = date.fromisoformat(date_str)
        except ValueError as e:
            raise HTTPException(
                status_code=400, detail="Invalid date format (use YYYY-MM-DD)"
            ) from e
    else:
        target_date = brisbane_today()

    # Load frozen predictions from S3
    try:
        frozen_rows = await load_frozen_watchlist(target_date)
    except Exception as e:
        logger.exception("Failed to load frozen watchlist from S3")
        raise HTTPException(status_code=500, detail=f"Failed to load predictions: {e}") from e

    if not frozen_rows:
        return WatchlistResponse(
            generated_at=datetime.now(UTC).isoformat(),
            entry_count=0,
            entries=[],
            message=f"No data available for {target_date}",
        )

    # Fetch live data from DB
    live_odds = await fetch_live_odds(session, target_date)
    results = await fetch_results(session, target_date)
    race_statuses = await fetch_race_statuses(session, target_date)

    # Fetch track states for all tracks in watchlist
    track_names = {row.get("Track") for row in frozen_rows if row.get("Track")}
    track_result = await session.execute(
        select(Track.name, Track.state).where(Track.name.in_(track_names))
    )
    track_states: dict[str, str | None] = {name: state for name, state in track_result.all()}

    # Fetch user bets if username provided
    bets_by_form_id: dict[tuple[int, str], UserBet] = {}
    if username:
        result = await session.execute(select(UserBet).where(UserBet.username == username))
        all_bets = result.scalars().all()
        bets_by_form_id = {
            (bet.form_id, bet.race_date.isoformat()): bet for bet in all_bets
        }

    # Fetch live paper bets from SQLite ledger
    race_ids = {int(row["Race ID"]) for row in frozen_rows if row.get("Race ID")}
    paper_bets = fetch_paper_bets(target_date, race_ids)

    # Compose entries
    entries: list[WatchlistEntry] = []
    for row in frozen_rows:
        try:
            race_id = int(row["Race ID"])
            form_id = int(row["Form ID"])
            horse_id = int(row["Horse ID"])

            # Frozen fields (from S3 artifact)
            our_win = float(row["Our Win"]) if row.get("Our Win") else 0.0
            our_place = float(row["Our Place"]) if row.get("Our Place") else 0.0
            win_trigger = float(row["WIN Trigger"]) if row.get("WIN Trigger") else 0.0
            place_trigger = float(row["PLACE Trigger"]) if row.get("PLACE Trigger") else 0.0
            in_monitor_net = row.get("In Monitor Net", "")  # FROZEN verdict

            # Live odds (keyed by race_id, horse_id)
            odds_key = (race_id, horse_id)
            odds = live_odds.get(odds_key, {})
            market_win = odds.get("fixed_win")
            market_place = odds.get("fixed_place")

            # Recompute overlays (live market × frozen fair)
            win_overlay_pct: float | None = None
            win_distance_to_trigger: float | None = None
            place_overlay_pct: float | None = None
            place_distance_to_trigger: float | None = None

            if market_win and our_win > 0:
                win_overlay_pct = ((market_win / our_win) - 1.0) * 100.0
                if win_trigger > 0:
                    win_distance_to_trigger = ((market_win - win_trigger) / win_trigger) * 100.0

            if market_place and our_place > 0:
                place_overlay_pct = ((market_place / our_place) - 1.0) * 100.0
                if place_trigger > 0:
                    place_distance_to_trigger = (
                        (market_place - place_trigger) / place_trigger
                    ) * 100.0

            # Live results (keyed by race_id, form_id)
            result_key = (race_id, form_id)
            result = results.get(result_key, {})
            actual_position = result.get("position")
            actual_margin = result.get("margin")
            is_scratched = result.get("is_scratched", False)

            # Race status
            race_status = race_statuses.get(race_id)

            # User bet (keyed by form_id, race_date)
            bet = bets_by_form_id.get((form_id, str(target_date)))
            bet_placed = bet is not None

            # Paper bets (keyed by race_id, horse_id)
            paper_bet = paper_bets.get((race_id, horse_id), {})

            # Market rank (compute from live odds ordering within race)
            market_rank: int | None = None
            if market_win:
                # Collect all runners in this race with live odds
                race_odds = [
                    (hid, live_odds[(rid, hid)]["fixed_win"])
                    for (rid, hid) in live_odds
                    if rid == race_id and live_odds[(rid, hid)].get("fixed_win")
                ]
                race_odds_sorted = sorted(race_odds, key=lambda x: x[1] if x[1] else 999.0)
                market_rank = next(
                    (i + 1 for i, (hid, _) in enumerate(race_odds_sorted) if hid == horse_id),
                    None,
                )

            track_name = row.get("Track", "")
            entries.append(
                WatchlistEntry(
                    track=track_name,
                    country=row.get("Country", "AUS"),
                    track_state=track_states.get(track_name),
                    race_number=int(row.get("Race #", 0)) if row.get("Race #") else 0,
                    race_time=row.get("Race Time", ""),
                    race_name=row.get("Race Name", ""),
                    distance_m=int(row.get("Distance", 0)) if row.get("Distance") else 0,
                    class_description=row.get("Class", ""),
                    track_condition=row.get("Track Condition", ""),
                    horse=row.get("Horse", ""),
                    jockey_name=row.get("Jockey") or None,
                    weight_kg=float(row["Weight"]) if row.get("Weight") else None,
                    barrier=int(row["Barrier"]) if row.get("Barrier") else None,
                    our_win=our_win,
                    win_pct=float(row["Win %"]) if row.get("Win %") else 0.0,
                    win_trigger=win_trigger,
                    our_place=our_place,
                    place_pct=float(row["Place %"]) if row.get("Place %") else 0.0,
                    place_trigger=place_trigger,
                    win_overlay_pct=(
                        round(win_overlay_pct, 1) if win_overlay_pct is not None else None
                    ),
                    win_distance_to_trigger=(
                        round(win_distance_to_trigger, 1)
                        if win_distance_to_trigger is not None
                        else None
                    ),
                    market_win=round(market_win, 2) if market_win else None,
                    market_place=round(market_place, 2) if market_place else None,
                    market_rank=market_rank,
                    place_overlay_pct=(
                        round(place_overlay_pct, 1) if place_overlay_pct is not None else None
                    ),
                    place_distance_to_trigger=(
                        round(place_distance_to_trigger, 1)
                        if place_distance_to_trigger is not None
                        else None
                    ),
                    neds_win=float(row["Neds Win"]) if row.get("Neds Win") else None,
                    neds_place=float(row["Neds Place"]) if row.get("Neds Place") else None,
                    class_rank=int(row["Class Rank"]) if row.get("Class Rank") else 0,
                    sim_order=int(row["Sim Order"]) if row.get("Sim Order") else None,
                    sim_win_pct=float(row["Sim Win %"]) if row.get("Sim Win %") else None,
                    sim_win_rank=int(row["Sim Win Rank"]) if row.get("Sim Win Rank") else None,
                    sim_place_pct=float(row["Sim Place %"]) if row.get("Sim Place %") else None,
                    sim_place_rank=(
                        int(row["Sim Place Rank"]) if row.get("Sim Place Rank") else None
                    ),
                    ml_win_pct=float(row["ML Win %"]) if row.get("ML Win %") else None,
                    ml_win_rank=int(row["ML Win Rank"]) if row.get("ML Win Rank") else None,
                    ml_place_pct=float(row["ML Place %"]) if row.get("ML Place %") else None,
                    ml_place_rank=(
                        int(row["ML Place Rank"]) if row.get("ML Place Rank") else None
                    ),
                    pf_time_rank=int(row["PF Time Rank"]) if row.get("PF Time Rank") else None,
                    neds_url=row.get("Neds Link URL"),
                    in_monitor_net=in_monitor_net,  # FROZEN — monitor's actual decision
                    placed="",  # NULL — frozen PLACED is stale (morning builder, pre-settle)
                    race_id=race_id,
                    race_date=str(target_date),
                    form_id=form_id,
                    horse_id=horse_id,
                    actual_position=actual_position,
                    actual_margin=actual_margin,
                    is_scratched=is_scratched,
                    race_status=race_status,
                    bet_placed=bet_placed,
                    bet_id=bet.id if bet else None,
                    bet_type=bet.bet_type if bet else None,
                    odds_taken=float(bet.odds_taken) if bet and bet.odds_taken else None,
                    stake_aud=float(bet.stake_aud) if bet and bet.stake_aud else None,
                    result_position=bet.result_position if bet else None,
                    payout_aud=float(bet.payout_aud) if bet and bet.payout_aud else None,
                    profit_aud=float(bet.profit_aud) if bet and bet.profit_aud else None,
                    roi_pct=(
                        float((bet.profit_aud / bet.stake_aud) * 100)
                        if bet and bet.profit_aud and bet.stake_aud
                        else None
                    ),
                    paper_win_placed=paper_bet.get("win_placed", False),
                    paper_place_placed=paper_bet.get("place_placed", False),
                    paper_win_result=paper_bet.get("win_result"),
                    paper_place_result=paper_bet.get("place_result"),
                    paper_win_stake=paper_bet.get("win_stake"),
                    paper_place_stake=paper_bet.get("place_stake"),
                    paper_win_odds=paper_bet.get("win_odds"),
                    paper_place_odds=paper_bet.get("place_odds"),
                )
            )
        except (KeyError, ValueError) as e:
            logger.warning("Failed to compose row: %s — error: %s", row, e)
            continue

    logger.info(
        "GET /watchlist date=%s: %d entries composed (frozen + live)",
        target_date,
        len(entries),
    )

    return WatchlistResponse(
        generated_at=datetime.now(UTC).isoformat(),
        entry_count=len(entries),
        entries=entries,
        message=None if entries else f"No runners found for {target_date}",
    )


@router.get("/download")
async def download_watchlist(
    date_str: str | None = Query(
        None, alias="date", description="Target date YYYY-MM-DD (default Brisbane today)"
    ),
    session: AsyncSession = Depends(get_db_session),
) -> StreamingResponse:
    """Download watchlist as xlsx with live 3-source composition."""
    # Parse target date
    if date_str:
        try:
            target_date = date.fromisoformat(date_str)
        except ValueError as e:
            raise HTTPException(
                status_code=400, detail="Invalid date format (use YYYY-MM-DD)"
            ) from e
    else:
        target_date = brisbane_today()

    # Load frozen predictions from S3
    try:
        frozen_rows = await load_frozen_watchlist(target_date)
    except Exception as e:
        logger.exception("Failed to load frozen watchlist from S3")
        raise HTTPException(status_code=503, detail=f"Watchlist data unavailable: {e}") from e

    if not frozen_rows:
        raise HTTPException(status_code=404, detail=f"No data available for {target_date}")

    # Fetch live data
    live_odds = await fetch_live_odds(session, target_date)
    results = await fetch_results(session, target_date)

    # Build xlsx
    wb = Workbook()
    ws: Worksheet = wb.active  # type: ignore[assignment]
    ws.title = "Watchlist"

    headers = [
        "Track",
        "Country",
        "Race #",
        "Race Time",
        "Horse",
        "Our Win",
        "Win %",
        "Market Win",
        "Our Place",
        "Place %",
        "Market Place",
        "Actual Position",
        "Actual Margin",
        "Form ID",
        "Race Date",
        "Race ID",
        "Horse ID",
    ]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)  # type: ignore[misc]

    for row in frozen_rows:
        try:
            race_id = int(row["Race ID"])
            horse_id = int(row["Horse ID"])
            form_id = int(row["Form ID"])

            # Live odds
            odds_key = (race_id, horse_id)
            odds = live_odds.get(odds_key, {})
            market_win = odds.get("fixed_win")
            market_place = odds.get("fixed_place")

            # Live results
            result_key = (race_id, form_id)
            result = results.get(result_key, {})
            actual_position = result.get("position")
            actual_margin = result.get("margin")

            ws.append(
                [
                    row.get("Track", ""),
                    row.get("Country", "AUS"),
                    int(row["Race #"]) if row.get("Race #") else "",
                    row.get("Race Time", ""),
                    row.get("Horse", ""),
                    float(row["Our Win"]) if row.get("Our Win") else "",
                    float(row["Win %"]) if row.get("Win %") else "",
                    round(market_win, 2) if market_win else "",
                    float(row["Our Place"]) if row.get("Our Place") else "",
                    float(row["Place %"]) if row.get("Place %") else "",
                    round(market_place, 2) if market_place else "",
                    actual_position if actual_position else "",
                    round(actual_margin, 2) if actual_margin else "",
                    form_id,
                    str(target_date),
                    race_id,
                    horse_id,
                ]
            )
        except (KeyError, ValueError) as e:
            logger.warning("Failed to process row for download: %s — error: %s", row, e)
            continue

    # Set column widths
    for i in range(1, len(headers) - 3):
        ws.column_dimensions[get_column_letter(i)].width = 15

    # Hide ID columns
    for col_idx in [14, 15, 16, 17]:  # Form ID, Race Date, Race ID, Horse ID
        ws.column_dimensions[get_column_letter(col_idx)].hidden = True

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    filename = f"watchlist-{target_date}.xlsx"

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
