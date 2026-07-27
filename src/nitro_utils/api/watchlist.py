import csv
import io
import logging
import secrets
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import HTMLResponse, ORJSONResponse, StreamingResponse
from kubernetes import client, config  # type: ignore[import-untyped]
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession

from nitro_utils.config import settings
from nitro_utils.database import get_db_session
from nitro_utils.models import UserBet
from decimal import Decimal
from datetime import date as Date

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/watchlist", tags=["watchlist"])


class WatchlistEntry(BaseModel):
    track: str
    race_number: int
    race_time: str
    horse: str
    our_win: float
    win_pct: float
    win_trigger: float
    our_place: float
    place_pct: float
    place_trigger: float
    win_overlay_pct: float | None
    win_distance_to_trigger: float | None
    market_win: float | None
    market_place: float | None
    market_rank: int | None
    place_overlay_pct: float | None
    place_distance_to_trigger: float | None
    neds_win: float | None
    neds_place: float | None
    class_rank: int
    sim_order: int | None
    sim_win_pct: float | None
    sim_win_rank: int | None
    sim_place_pct: float | None
    sim_place_rank: int | None
    ml_win_pct: float | None
    ml_win_rank: int | None
    ml_place_pct: float | None
    ml_place_rank: int | None
    pf_time_rank: int | None
    neds_url: str | None
    in_monitor_net: str
    placed: str
    race_id: int
    race_date: str
    form_id: int
    horse_id: int
    bet_placed: bool
    bet_id: int | None
    bet_type: str | None
    odds_taken: float | None
    stake_aud: float | None
    result_position: int | None
    payout_aud: float | None
    profit_aud: float | None
    roi_pct: float | None


class BetSummary(BaseModel):
    total_bets: int
    total_stake_aud: float
    total_payout_aud: float
    total_profit_aud: float
    roi_pct: float


class WatchlistResponse(BaseModel):
    generated_at: str
    entry_count: int
    entries: list[WatchlistEntry]
    summary: BetSummary


def _parse_watchlist_csv_raw(csv_path: Path) -> list[dict[str, str]]:
    """Parse watchlist CSV into raw dict rows."""
    if not csv_path.exists():
        raise FileNotFoundError(f"Watchlist CSV not found at {csv_path}")

    rows: list[dict[str, str]] = []
    with csv_path.open("r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)
    return rows


def _render_html_table(entries: list[WatchlistEntry]) -> str:
    rows_html = ""
    for entry in entries:
        rows_html += f"""
        <tr>
            <td>{entry.track}</td>
            <td>{entry.race_number}</td>
            <td>{entry.race_time}</td>
            <td>{entry.horse}</td>
            <td>{entry.our_win:.2f}</td>
            <td>{entry.win_pct:.1f}%</td>
            <td>{entry.win_overlay_pct:.1f if entry.win_overlay_pct is not None else '-'}%</td>
            <td>{entry.win_distance_to_trigger:.2f if entry.win_distance_to_trigger is not None else '-'}</td>
            <td>{entry.market_win:.2f if entry.market_win is not None else '-'}</td>
            <td>{entry.placed}</td>
        </tr>
        """

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Watchlist - Nitro Wagering</title>
    <style>
        body {{ font-family: system-ui, -apple-system, sans-serif; margin: 20px; }}
        table {{ border-collapse: collapse; width: 100%; font-size: 14px; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        th {{ background-color: #f2f2f2; position: sticky; top: 0; }}
        tr:nth-child(even) {{ background-color: #f9f9f9; }}
        .header {{ margin-bottom: 20px; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>Watchlist</h1>
        <p>Last updated: {datetime.now(timezone.utc).isoformat()}</p>
        <p>Total entries: {len(entries)}</p>
    </div>
    <table>
        <thead>
            <tr>
                <th>Track</th>
                <th>Race</th>
                <th>Time</th>
                <th>Horse</th>
                <th>Our Win</th>
                <th>Win %</th>
                <th>Overlay %</th>
                <th>Distance to Trigger</th>
                <th>Market Win</th>
                <th>Placed</th>
            </tr>
        </thead>
        <tbody>
            {rows_html}
        </tbody>
    </table>
</body>
</html>"""


@router.get("", response_model=WatchlistResponse)
async def get_watchlist(username: str, session: AsyncSession = Depends(get_db_session)) -> WatchlistResponse:
    csv_path = Path(settings.watchlist_csv_path)

    try:
        csv_rows = _parse_watchlist_csv_raw(csv_path)
    except FileNotFoundError:
        raise HTTPException(status_code=503, detail="Watchlist data unavailable")
    except Exception as e:
        logger.exception("Failed to parse watchlist CSV")
        raise HTTPException(status_code=500, detail=f"Failed to load watchlist: {e}")

    form_ids = [int(row["Form ID"]) for row in csv_rows if row.get("Form ID")]
    csv_keys = {(int(row["Form ID"]), row["Race Date"]) for row in csv_rows if row.get("Form ID")}

    result = await session.execute(
        select(UserBet).where(UserBet.username == username)
    )
    all_bets = result.scalars().all()
    bets_by_form_id: dict[tuple[int, str], UserBet] = {
        (bet.form_id, bet.race_date.isoformat()): bet for bet in all_bets
    }

    matched_bet_keys = set()
    entries: list[WatchlistEntry] = []
    total_stake = 0.0
    total_payout = 0.0
    total_profit = 0.0
    bet_count = 0

    for row in csv_rows:
        try:
            form_id = int(row["Form ID"])
            race_date = row["Race Date"]
            bet = bets_by_form_id.get((form_id, race_date))

            bet_placed = bet is not None
            if bet:
                matched_bet_keys.add((form_id, race_date))
                bet_count += 1
                total_stake += float(bet.stake_aud)
                if bet.payout_aud:
                    total_payout += float(bet.payout_aud)
                if bet.profit_aud:
                    total_profit += float(bet.profit_aud)

            entries.append(
                WatchlistEntry(
                    track=row["Track"],
                    race_number=int(row["Race #"]),
                    race_time=row["Race Time"],
                    horse=row["Horse"],
                    our_win=float(row["Our Win"]),
                    win_pct=float(row["Win %"]),
                    win_trigger=float(row["WIN Trigger"]),
                    our_place=float(row["Our Place"]),
                    place_pct=float(row["Place %"]),
                    place_trigger=float(row["PLACE Trigger"]),
                    win_overlay_pct=float(row["Win Overlay %"]) if row.get("Win Overlay %") else None,
                    win_distance_to_trigger=float(row["Win Distance to Trigger"]) if row.get("Win Distance to Trigger") else None,
                    market_win=float(row["Market Win"]) if row.get("Market Win") else None,
                    market_place=float(row["Market Place"]) if row.get("Market Place") else None,
                    market_rank=int(row["Market Rank"]) if row.get("Market Rank") else None,
                    place_overlay_pct=float(row["Place Overlay %"]) if row.get("Place Overlay %") else None,
                    place_distance_to_trigger=float(row["Place Distance to Trigger"]) if row.get("Place Distance to Trigger") else None,
                    neds_win=float(row["Neds Win"]) if row.get("Neds Win") else None,
                    neds_place=float(row["Neds Place"]) if row.get("Neds Place") else None,
                    class_rank=int(row["Class Rank"]),
                    sim_order=int(row["Sim Order"]) if row.get("Sim Order") else None,
                    sim_win_pct=float(row["Sim Win %"]) if row.get("Sim Win %") else None,
                    sim_win_rank=int(row["Sim Win Rank"]) if row.get("Sim Win Rank") else None,
                    sim_place_pct=float(row["Sim Place %"]) if row.get("Sim Place %") else None,
                    sim_place_rank=int(row["Sim Place Rank"]) if row.get("Sim Place Rank") else None,
                    ml_win_pct=float(row["ML Win %"]) if row.get("ML Win %") else None,
                    ml_win_rank=int(row["ML Win Rank"]) if row.get("ML Win Rank") else None,
                    ml_place_pct=float(row["ML Place %"]) if row.get("ML Place %") else None,
                    ml_place_rank=int(row["ML Place Rank"]) if row.get("ML Place Rank") else None,
                    pf_time_rank=int(row["PF Time Rank"]) if row.get("PF Time Rank") else None,
                    neds_url=row.get("Neds Link URL"),
                    in_monitor_net=row["In Monitor Net"],
                    placed=row["PLACED"],
                    race_id=int(row["Race ID"]),
                    race_date=race_date,
                    form_id=form_id,
                    horse_id=int(row["Horse ID"]),
                    bet_placed=bet_placed,
                    bet_id=bet.id if bet else None,
                    bet_type=bet.bet_type if bet else None,
                    odds_taken=float(bet.odds_taken) if bet and bet.odds_taken else None,
                    stake_aud=float(bet.stake_aud) if bet and bet.stake_aud else None,
                    result_position=bet.result_position if bet else None,
                    payout_aud=float(bet.payout_aud) if bet and bet.payout_aud else None,
                    profit_aud=float(bet.profit_aud) if bet and bet.profit_aud else None,
                    roi_pct=float((bet.profit_aud / bet.stake_aud) * 100) if bet and bet.profit_aud and bet.stake_aud else None,
                )
            )
        except (KeyError, ValueError) as e:
            logger.warning("Failed to parse CSV row: %s — error: %s", row, e)
            continue

    unmatched_bets = [
        bet for bet in all_bets
        if (bet.form_id, bet.race_date.isoformat()) not in matched_bet_keys
    ]

    for bet in unmatched_bets:
        bet_count += 1
        total_stake += float(bet.stake_aud)
        if bet.payout_aud:
            total_payout += float(bet.payout_aud)
        if bet.profit_aud:
            total_profit += float(bet.profit_aud)

        entries.append(
            WatchlistEntry(
                track="Unknown",
                race_number=0,
                race_time="",
                horse="Unknown",
                our_win=0.0,
                win_pct=0.0,
                win_trigger=0.0,
                our_place=0.0,
                place_pct=0.0,
                place_trigger=0.0,
                win_overlay_pct=None,
                win_distance_to_trigger=None,
                market_win=None,
                market_place=None,
                market_rank=None,
                place_overlay_pct=None,
                place_distance_to_trigger=None,
                neds_win=None,
                neds_place=None,
                class_rank=0,
                sim_order=None,
                sim_win_pct=None,
                sim_win_rank=None,
                sim_place_pct=None,
                sim_place_rank=None,
                ml_win_pct=None,
                ml_win_rank=None,
                ml_place_pct=None,
                ml_place_rank=None,
                pf_time_rank=None,
                neds_url=None,
                in_monitor_net="",
                placed="",
                race_id=0,
                race_date=bet.race_date.isoformat(),
                form_id=bet.form_id,
                horse_id=0,
                bet_placed=True,
                bet_id=bet.id,
                bet_type=bet.bet_type,
                odds_taken=float(bet.odds_taken) if bet.odds_taken else None,
                stake_aud=float(bet.stake_aud) if bet.stake_aud else None,
                result_position=bet.result_position,
                payout_aud=float(bet.payout_aud) if bet.payout_aud else None,
                profit_aud=float(bet.profit_aud) if bet.profit_aud else None,
                roi_pct=float((bet.profit_aud / bet.stake_aud) * 100) if bet.profit_aud and bet.stake_aud else None,
            )
        )

    logger.info(
        "GET /watchlist: username=%s total_entries=%d matched=%d unmatched=%d",
        username,
        len(entries),
        len(entries) - len(unmatched_bets),
        len(unmatched_bets),
    )

    stat = csv_path.stat()
    generated_at = datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat()

    roi_pct = (total_profit / total_stake * 100) if total_stake > 0 else 0.0

    return WatchlistResponse(
        generated_at=generated_at,
        entry_count=len(entries),
        entries=entries,
        summary=BetSummary(
            total_bets=bet_count,
            total_stake_aud=round(total_stake, 2),
            total_payout_aud=round(total_payout, 2),
            total_profit_aud=round(total_profit, 2),
            roi_pct=round(roi_pct, 2),
        ),
    )


@router.post("/refresh")
async def refresh_watchlist() -> dict[str, Any]:
    try:
        config.load_incluster_config()
    except config.ConfigException:
        try:
            config.load_kube_config()
        except config.ConfigException:
            raise HTTPException(status_code=500, detail="Kubernetes config unavailable")

    batch_v1 = client.BatchV1Api()
    job_name = f"watchlist-refresh-{secrets.token_hex(4)}"

    job_manifest = {
        "apiVersion": "batch/v1",
        "kind": "Job",
        "metadata": {"name": job_name, "namespace": settings.k8s_namespace},
        "spec": {
            "template": {
                "spec": {
                    "containers": [
                        {
                            "name": "refresh",
                            "image": settings.k8s_job_image,
                            "command": ["python3", "/tmp/build_watchlist_final.py"],
                            "env": [
                                {
                                    "name": "NITRO_DATABASE_URL",
                                    "valueFrom": {
                                        "secretKeyRef": {
                                            "name": "nitro-secrets",
                                            "key": "database-url",
                                        }
                                    },
                                }
                            ],
                            "volumeMounts": [
                                {"name": "shared", "mountPath": "/shared"},
                                {"name": "workspace", "mountPath": "/workspace"},
                            ],
                        }
                    ],
                    "volumes": [
                        {"name": "shared", "persistentVolumeClaim": {"claimName": "nitro-shared"}},
                        {
                            "name": "workspace",
                            "persistentVolumeClaim": {"claimName": "paper-monitor-workspace"},
                        },
                    ],
                    "restartPolicy": "OnFailure",
                }
            },
            "backoffLimit": 2,
        },
    }

    try:
        batch_v1.create_namespaced_job(namespace=settings.k8s_namespace, body=job_manifest)
        logger.info("Created watchlist refresh job: %s", job_name)
        return {"status": "triggered", "job_name": job_name}
    except Exception as e:
        logger.exception("Failed to create watchlist refresh job")
        raise HTTPException(status_code=500, detail=f"Failed to trigger refresh: {e}")


@router.get("/download")
async def download_watchlist() -> StreamingResponse:
    csv_path = Path(settings.watchlist_csv_path)

    try:
        csv_rows = _parse_watchlist_csv_raw(csv_path)
    except FileNotFoundError:
        raise HTTPException(status_code=503, detail="Watchlist data unavailable")
    except Exception as e:
        logger.exception("Failed to parse watchlist CSV")
        raise HTTPException(status_code=500, detail=f"Failed to load watchlist: {e}")

    wb = Workbook()
    ws: Worksheet = wb.active  # type: ignore[assignment]
    ws.title = "Watchlist"

    headers = [
        "Track",
        "Country",
        "Race #",
        "Race Time",
        "Horse",
        "Our Win",
        "Win %",
        "Neds Win",
        "Our Place",
        "Place %",
        "Neds Place",
        "Bet Placed",
        "Bet Type",
        "Odds Taken",
        "Stake (AUD)",
        "Result Position",
        "Payout (AUD)",
        "Profit (AUD)",
        "ROI %",
        "Form ID",
        "Race Date",
        "Race ID",
        "Horse ID",
    ]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)  # type: ignore[misc]

    for row in csv_rows:
        try:
            ws.append(
                [
                    row["Track"],
                    "AUS",  # TODO: extract country from data
                    int(row["Race #"]),
                    row["Race Time"],
                    row["Horse"],
                    float(row["Our Win"]),
                    float(row["Win %"]),
                    float(row["Neds Win"]) if row.get("Neds Win") else "",
                    float(row["Our Place"]),
                    float(row["Place %"]),
                    float(row["Neds Place"]) if row.get("Neds Place") else "",
                    "",  # Bet Placed (user fills)
                    "",  # Bet Type (user fills)
                    "",  # Odds Taken (user fills)
                    "",  # Stake (user fills)
                    "",  # Result Position (computed post-race)
                    "",  # Payout (computed post-race)
                    "",  # Profit (computed post-race)
                    "",  # ROI % (computed post-race)
                    int(row["Form ID"]),  # Hidden ID column for upload
                    row["Race Date"],  # Hidden ID column for upload
                    int(row["Race ID"]),  # Hidden ID column for upload
                    int(row["Horse ID"]),  # Hidden ID column for upload
                ]
            )
        except (KeyError, ValueError) as e:
            logger.warning("Failed to process row for download: %s — error: %s", row, e)
            continue

    for i in range(1, len(headers) - 3):
        ws.column_dimensions[get_column_letter(i)].width = 15

    ws.column_dimensions[get_column_letter(20)].hidden = True
    ws.column_dimensions[get_column_letter(21)].hidden = True
    ws.column_dimensions[get_column_letter(22)].hidden = True
    ws.column_dimensions[get_column_letter(23)].hidden = True

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    filename = f"watchlist-{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


class UploadResult(BaseModel):
    status: str
    imported_bets: int
    errors: list[str]


@router.post("/upload", response_model=UploadResult)
async def upload_watchlist(username: str, file: UploadFile = File(...), session: AsyncSession = Depends(get_db_session)) -> UploadResult:
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="File must be .xlsx format")

    try:
        from openpyxl import load_workbook
        from sqlalchemy.exc import IntegrityError
        from sqlalchemy import text

        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active

        headers_row = [cell.value for cell in ws[1]]  # type: ignore[index]
        expected_headers_base = [
            "Track",
            "Country",
            "Race #",
            "Race Time",
            "Horse",
            "Our Win",
            "Win %",
            "Neds Win",
            "Our Place",
            "Place %",
            "Neds Place",
            "Bet Placed",
            "Bet Type",
            "Odds Taken",
            "Stake (AUD)",
            "Result Position",
            "Payout (AUD)",
            "Profit (AUD)",
            "ROI %",
        ]

        expected_id_headers = ["Form ID", "Race Date", "Race ID", "Horse ID"]
        expected_full = expected_headers_base + expected_id_headers

        if headers_row[:19] != expected_headers_base:
            raise HTTPException(
                status_code=400,
                detail="Invalid file format. Please download the template from this service (GET /download).",
            )

        if len(headers_row) < 23 or headers_row[19:23] != expected_id_headers:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required ID columns (Form ID, Race Date, Race ID, Horse ID). Please download the template from this service (GET /download).",
            )

        await session.execute(
            text("INSERT INTO tracker_users (username) VALUES (:username) ON CONFLICT (username) DO NOTHING"),
            {"username": username},
        )

        imported_count = 0
        errors: list[str] = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):  # type: ignore[arg-type]
            try:
                bet_placed = str(row[11]).strip().upper() if row[11] else ""
                if bet_placed not in ("YES", "Y", "TRUE", "1"):
                    continue

                bet_type = str(row[12]).strip().lower() if row[12] else None
                odds_taken = float(row[13]) if row[13] else None
                stake = float(row[14]) if row[14] else None

                form_id = int(row[19]) if len(row) > 19 and row[19] else None
                race_date_str = str(row[20]) if len(row) > 20 and row[20] else None

                if not form_id or not race_date_str:
                    errors.append(f"Row {row_idx}: Missing Form ID or Race Date")
                    continue

                if not all([bet_type, odds_taken, stake]):
                    errors.append(f"Row {row_idx}: Missing required bet fields")
                    continue

                if bet_type not in ("win", "place", "each_way"):
                    errors.append(f"Row {row_idx}: Invalid bet type '{bet_type}'")
                    continue

                if odds_taken <= 0 or stake <= 0:  # type: ignore[operator]
                    errors.append(f"Row {row_idx}: Odds and stake must be positive")
                    continue

                race_date = Date.fromisoformat(race_date_str)

                new_bet = UserBet(
                    username=username,
                    form_id=form_id,
                    race_date=race_date,
                    bet_type=bet_type,
                    odds_taken=Decimal(str(odds_taken)),
                    stake_aud=Decimal(str(stake)),
                )

                try:
                    session.add(new_bet)
                    await session.flush()
                    imported_count += 1
                except IntegrityError as ie:
                    await session.rollback()
                    errors.append(f"Row {row_idx}: Duplicate bet - {ie}")
                    continue

            except (ValueError, TypeError) as e:
                errors.append(f"Row {row_idx}: {type(e).__name__} - {e}")
                continue

        await session.commit()

        logger.info("Upload processing complete: %d bets imported, %d errors", imported_count, len(errors))

        return UploadResult(
            status="ok",
            imported_bets=imported_count,
            errors=errors[:20],
        )

    except Exception as e:
        await session.rollback()
        logger.exception("Failed to process uploaded file")
        raise HTTPException(status_code=500, detail=f"Failed to process file: {e}")
